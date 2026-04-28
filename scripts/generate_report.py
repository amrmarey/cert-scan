#!/usr/bin/env python3
"""Generate a styled XLSX certificate inventory report from scan JSON data."""

import json
import socket
import sys
from datetime import datetime, timezone

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

HEADERS = [
    "Asset_IP_Add",
    "Hostname",
    "Port",
    "HTTPS",
    "Cert_Issuer",
    "Serial_Number",
    "Serial_Hex",
    "Subject_Alt_Names",
    "Days_Remaining",
    "Expiry_Date",
    "Expiry_Status",
    "Error",
]

COL_WIDTHS = {
    "Asset_IP_Add":      22,
    "Hostname":          30,
    "Port":               8,
    "HTTPS":              8,
    "Cert_Issuer":       48,
    "Serial_Number":     32,
    "Serial_Hex":        30,
    "Subject_Alt_Names": 58,
    "Days_Remaining":    16,
    "Expiry_Date":       24,
    "Expiry_Status":     16,
    "Error":             42,
}

STATUS_STYLES = {
    "Expired":  {"fill": "C00000", "font_color": "FFFFFF"},
    "Critical": {"fill": "FF6600", "font_color": "FFFFFF"},
    "Warning":  {"fill": "FFCC00", "font_color": "000000"},
    "Healthy":  {"fill": "00B050", "font_color": "FFFFFF"},
    "N/A":      {"fill": "BFBFBF", "font_color": "000000"},
}

_DNS_TIMEOUT = 2  # seconds per reverse-DNS lookup


def _is_ip(address):
    for family in (socket.AF_INET, socket.AF_INET6):
        try:
            socket.inet_pton(family, address)
            return True
        except socket.error:
            pass
    return False


def reverse_dns(address):
    """Return PTR hostname for an IP address; empty string for hostnames or failures."""
    if not _is_ip(address):
        return ""
    old_timeout = socket.getdefaulttimeout()
    try:
        socket.setdefaulttimeout(_DNS_TIMEOUT)
        return socket.gethostbyaddr(address)[0]
    except Exception:
        return ""
    finally:
        socket.setdefaulttimeout(old_timeout)


def parse_asn1_date(s):
    if not s:
        return None
    s = s.strip()
    if s.endswith("Z"):
        return datetime.strptime(s, "%Y%m%d%H%M%SZ").replace(tzinfo=timezone.utc)
    return datetime.strptime(s[:14], "%Y%m%d%H%M%S").replace(tzinfo=timezone.utc)


def days_remaining(date_str):
    dt = parse_asn1_date(date_str)
    if dt is None:
        return None
    return (dt - datetime.now(timezone.utc)).days


def expiry_status(days):
    if days is None:
        return "N/A"
    if days < 0:
        return "Expired"
    if days < 30:
        return "Critical"
    if days < 90:
        return "Warning"
    return "Healthy"


def format_dn(issuer_dict):
    return ", ".join(f"{k}={v}" for k, v in issuer_dict.items())


def format_expiry_date(date_str):
    dt = parse_asn1_date(date_str)
    return dt.strftime("%Y-%m-%d %H:%M UTC") if dt else "N/A"


def format_serial_hex(serial_int):
    if not serial_int:
        return "N/A"
    return f"0x{serial_int:X}"


def build_row(entry):
    address = entry["address"]
    hostname = reverse_dns(address)

    if entry.get("failed", True):
        return {
            "Asset_IP_Add":      address,
            "Hostname":          hostname,
            "Port":              str(entry["port"]),
            "HTTPS":             "No",
            "Cert_Issuer":       "N/A",
            "Serial_Number":     "N/A",
            "Serial_Hex":        "N/A",
            "Subject_Alt_Names": "N/A",
            "Days_Remaining":    "N/A",
            "Expiry_Date":       "N/A",
            "Expiry_Status":     "N/A",
            "Error":             entry.get("msg", "Connection failed"),
        }

    days = days_remaining(entry.get("not_after", ""))
    sans = entry.get("subject_alt_name") or []
    serial_int = entry.get("serial_number") or 0

    return {
        "Asset_IP_Add":      address,
        "Hostname":          hostname,
        "Port":              str(entry["port"]),
        "HTTPS":             "Yes",
        "Cert_Issuer":       format_dn(entry.get("issuer") or {}),
        "Serial_Number":     str(serial_int) if serial_int else "N/A",
        "Serial_Hex":        format_serial_hex(serial_int),
        "Subject_Alt_Names": "; ".join(sans),
        "Days_Remaining":    str(days) if days is not None else "N/A",
        "Expiry_Date":       format_expiry_date(entry.get("not_after", "")),
        "Expiry_Status":     expiry_status(days),
        "Error":             "",
    }


def write_xlsx(rows, output_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Certificate Inventory"

    thin = Side(style="thin", color="D9D9D9")
    cell_border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Header row
    ws.append(HEADERS)
    ws.row_dimensions[1].height = 22
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
    for col_idx, h in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = cell_border
        ws.column_dimensions[get_column_letter(col_idx)].width = COL_WIDTHS.get(h, 20)

    status_col = HEADERS.index("Expiry_Status") + 1
    san_col    = HEADERS.index("Subject_Alt_Names") + 1

    # Data rows
    for row_idx, row in enumerate(rows, start=2):
        ws.append([row.get(h, "") for h in HEADERS])
        ws.row_dimensions[row_idx].height = 16

        status = row.get("Expiry_Status", "N/A")
        style = STATUS_STYLES.get(status, STATUS_STYLES["N/A"])
        status_fill = PatternFill(
            start_color=style["fill"], end_color=style["fill"], fill_type="solid"
        )
        status_font = Font(bold=True, color=style["font_color"], name="Calibri", size=10)
        data_font = Font(name="Calibri", size=10)

        for col_idx in range(1, len(HEADERS) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = cell_border
            cell.alignment = Alignment(
                vertical="center",
                wrap_text=(col_idx == san_col),
            )
            if col_idx == status_col:
                cell.fill = status_fill
                cell.font = status_font
            else:
                cell.font = data_font

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    wb.save(output_path)


def print_summary(rows):
    print(
        f"\n{'Asset':<32} {'Hostname':<28} {'Port':<6} {'Status':<10} {'Days':>6}  Expiry"
    )
    print("-" * 96)
    for row in rows:
        hostname = row["Hostname"] or "-"
        print(
            f"{row['Asset_IP_Add']:<32} {hostname:<28} {row['Port']:<6}"
            f" {row['Expiry_Status']:<10} {row['Days_Remaining']:>6}  {row['Expiry_Date']}"
        )
    print()


def main():
    if len(sys.argv) < 3:
        print(f"Usage: {sys.argv[0]} <scan_data.json> <output.xlsx>", file=sys.stderr)
        sys.exit(1)

    json_path, xlsx_path = sys.argv[1], sys.argv[2]

    with open(json_path) as f:
        scan_data = json.load(f)

    print(f"Running reverse DNS lookups for {sum(1 for e in scan_data if _is_ip(e['address']))} IP address(es)...")
    rows = [build_row(entry) for entry in scan_data]
    print_summary(rows)
    write_xlsx(rows, xlsx_path)
    print(f"Scanned {len(rows)} asset(s). Report saved to: {xlsx_path}")


if __name__ == "__main__":
    main()
